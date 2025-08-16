"""
Tests for the flagged-csv converter.
"""

import pytest
import tempfile
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from flagged_csv import XlsxConverter, XlsxConverterConfig


def create_test_excel(file_path: Path, with_colors=False, with_merge=False):
    """Create a test Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    
    # Add some data
    ws['A1'] = 'Header1'
    ws['B1'] = 'Header2'
    ws['C1'] = 'Header3'
    
    ws['A2'] = 'Value1'
    ws['B2'] = 100
    ws['C2'] = 200.50
    
    # Add colors if requested
    if with_colors:
        ws['A1'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        ws['B2'].fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    
    # Add merge if requested
    if with_merge:
        ws['D1'] = 'Merged Cell'
        ws.merge_cells('D1:F1')
    
    wb.save(file_path)
    return wb


class TestXlsxConverter:
    """Test the XLSX converter functionality."""
    
    def test_basic_conversion(self):
        """Test basic XLSX to CSV conversion."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create test file
            xlsx_path = Path(temp_dir) / "test.xlsx"
            create_test_excel(xlsx_path)
            
            # Convert
            converter = XlsxConverter()
            result = converter.convert_to_csv(str(xlsx_path), 'TestSheet')
            
            # Check result
            assert 'Header1' in result
            assert 'Value1' in result
            assert '100' in result
    
    def test_color_extraction(self):
        """Test color flag extraction."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create test file with colors
            xlsx_path = Path(temp_dir) / "test_colors.xlsx"
            create_test_excel(xlsx_path, with_colors=True)
            
            # Convert with colors
            converter = XlsxConverter()
            result = converter.convert_to_csv(
                str(xlsx_path), 
                'TestSheet',
                include_colors=True
            )
            
            # Check for color flags
            assert '{#FF0000}' in result  # Red color
            assert '{#00FF00}' in result  # Green color
    
    def test_merge_detection(self):
        """Test merge cell detection."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create test file with merged cells
            xlsx_path = Path(temp_dir) / "test_merge.xlsx"
            create_test_excel(xlsx_path, with_merge=True)
            
            # Convert with merge detection
            converter = XlsxConverter()
            result = converter.convert_to_csv(
                str(xlsx_path),
                'TestSheet',
                signal_merge=True
            )
            
            # Check for merge flags
            assert '{MG:' in result
            assert 'Merged Cell' in result
    
    def test_ignore_colors(self):
        """Test ignoring specific colors."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create test file
            xlsx_path = Path(temp_dir) / "test_ignore.xlsx"
            wb = Workbook()
            ws = wb.active
            
            ws['A1'] = 'White BG'
            ws['A1'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            ws['B1'] = 'Red BG'
            ws['B1'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            
            wb.save(xlsx_path)
            
            # Convert ignoring white
            converter = XlsxConverter()
            result = converter.convert_to_csv(
                str(xlsx_path),
                'Sheet',
                include_colors=True,
                ignore_colors='#FFFFFF'
            )
            
            # White should be ignored, red should be included
            assert '{#FFFFFF}' not in result
            assert '{#FF0000}' in result
    
    def test_invalid_sheet_name(self):
        """Test error handling for invalid sheet names."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create test file
            xlsx_path = Path(temp_dir) / "test.xlsx"
            create_test_excel(xlsx_path)
            
            # Try to convert non-existent sheet
            converter = XlsxConverter()
            with pytest.raises(ValueError, match="Sheet 'InvalidSheet' not found"):
                converter.convert_to_csv(str(xlsx_path), 'InvalidSheet')
    
    def test_file_not_found(self):
        """Test error handling for missing files."""
        converter = XlsxConverter()
        with pytest.raises(FileNotFoundError):
            converter.convert_to_csv('/nonexistent/file.xlsx', 'Sheet1')
    
    def test_configuration(self):
        """Test converter configuration options."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create test file
            xlsx_path = Path(temp_dir) / "test.xlsx"
            create_test_excel(xlsx_path)
            
            # Test with header=False (default)
            config = XlsxConverterConfig(
                header=False,
                keep_default_na=True
            )
            converter = XlsxConverter(config)
            result = converter.convert_to_csv(str(xlsx_path), 'TestSheet')
            
            # When header=False, we should not output DataFrame column headers (A, B, C)
            # but we should still include all rows of data
            lines = result.strip().split('\n')
            assert len(lines) == 2  # Both rows of data
            assert 'Header1' in lines[0]
            assert 'Value1' in lines[1]
            
            # Test with header=True
            config_with_header = XlsxConverterConfig(
                header=True,
                keep_default_na=True
            )
            converter_with_header = XlsxConverter(config_with_header)
            result_with_header = converter_with_header.convert_to_csv(str(xlsx_path), 'TestSheet')
            
            # When header=True, we should output DataFrame column headers (A, B, C)
            lines_with_header = result_with_header.strip().split('\n')
            assert len(lines_with_header) == 3  # Column headers + 2 data rows
            assert 'A,B,C' in lines_with_header[0]
            assert 'Header1' in lines_with_header[1]
    
    def test_output_formats(self):
        """Test different output formats."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create test file
            xlsx_path = Path(temp_dir) / "test.xlsx"
            create_test_excel(xlsx_path)
            
            converter = XlsxConverter()
            
            # Test HTML output
            html_result = converter.convert_to_csv(
                str(xlsx_path),
                'TestSheet',
                output_format='html'
            )
            assert '<table' in html_result
            assert 'Header1' in html_result  # Header is in <th> not <td>
            assert '<td>Value1</td>' in html_result
            
            # Test Markdown output
            md_result = converter.convert_to_csv(
                str(xlsx_path),
                'TestSheet',
                output_format='markdown'
            )
            assert '|' in md_result  # Markdown tables use pipes
    
    def test_add_location(self):
        """Test adding location coordinates to cells."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create test file
            xlsx_path = Path(temp_dir) / "test_location.xlsx"
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Header1'
            ws['B1'] = 'Header2'
            ws['A2'] = 'Value1'
            ws['B2'] = 100
            wb.save(xlsx_path)
            
            # Convert with location tags
            converter = XlsxConverter()
            result = converter.convert_to_csv(
                str(xlsx_path),
                'Sheet',
                add_location=True
            )
            
            # Check for location tags
            assert '{l:A1}' in result
            assert '{l:B1}' in result
            assert '{l:A2}' in result
            assert '{l:B2}' in result
    
    def test_keep_empty_lines(self):
        """Test keeping empty rows."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create test file with empty rows
            xlsx_path = Path(temp_dir) / "test_empty.xlsx"
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Row1'
            # Row 2 is empty
            ws['A3'] = 'Row3'
            # Row 4 is empty
            ws['A5'] = 'Row5'
            wb.save(xlsx_path)
            
            # Convert without keep_empty_lines (default)
            converter = XlsxConverter()
            result_no_empty = converter.convert_to_csv(str(xlsx_path), 'Sheet')
            lines_no_empty = result_no_empty.strip().split('\n')
            # Should have 3 data rows (no empty lines)
            assert len(lines_no_empty) == 3
            
            # Convert with keep_empty_lines
            result_with_empty = converter.convert_to_csv(
                str(xlsx_path), 'Sheet',
                keep_empty_lines=True
            )
            lines_with_empty = result_with_empty.strip().split('\n')
            # Should have 5 rows (including empty lines)
            assert len(lines_with_empty) == 5
    
    def test_black_text_contrast(self):
        """Test that black text is only included when there's a background color for contrast."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create test file with various text/background combinations
            xlsx_path = Path(temp_dir) / "test_contrast.xlsx"
            wb = Workbook()
            ws = wb.active
            
            # White text on dark background
            ws['A1'] = 'White on Dark'
            ws['A1'].font = Font(color='FFFFFF')
            ws['A1'].fill = PatternFill(start_color='8E1C02', end_color='8E1C02', fill_type='solid')
            
            # Black text on light background
            ws['B1'] = 'Black on Light'
            ws['B1'].font = Font(color='000000')
            ws['B1'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # Black text with no background (should NOT include fc:#000000)
            ws['A2'] = 'Plain Black'
            ws['A2'].font = Font(color='000000')
            
            # Colored text with no background (should include the color)
            ws['B2'] = 'Red Text'
            ws['B2'].font = Font(color='FF0000')
            
            wb.save(xlsx_path)
            
            # Test 1: With default ignore lists (black fg ignored by default)
            converter = XlsxConverter()
            result = converter.convert_to_csv(
                str(xlsx_path),
                'Sheet',
                include_colors=True  # Include both fg and bg
            )
            
            lines = result.strip().split('\n')
            
            # Check white on dark background
            assert '{#8E1C02}' in lines[0]  # Dark background
            assert '{fc:#FFFFFF}' in lines[0]  # White text
            
            # Check black on light background (black text is ignored by default)
            assert '{#FFFF00}' in lines[0]  # Yellow background
            assert '{fc:#000000}' not in lines[0]  # Black text ignored by default
            
            # Check plain black text (no background)
            assert 'Plain Black' in lines[1]
            assert '{fc:#000000}' not in lines[1]  # Black text ignored by default
            
            # Check colored text (no background)
            assert 'Red Text' in lines[1]
            assert '{fc:#FF0000}' in lines[1]  # Red text should be included
            
            # Test 2: Without ignoring black text (to test contrast logic)
            result2 = converter.convert_to_csv(
                str(xlsx_path),
                'Sheet',
                include_colors=True,
                ignore_fg_colors=''  # Don't ignore any foreground colors
            )
            
            lines2 = result2.strip().split('\n')
            
            # Now black text on background should be included
            assert '{#FFFF00}' in lines2[0]  # Yellow background
            assert '{fc:#000000}' in lines2[0]  # Black text now included with background
            
            # Plain black text (no background) should still not be included due to contrast logic
            assert 'Plain Black' in lines2[1]
            assert '{fc:#000000}' not in lines2[1]  # Black text without background still not included
    
    def test_ignore_colors_defaults(self):
        """Test default ignore colors and custom ignore lists."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create test file with various color combinations
            xlsx_path = Path(temp_dir) / "test_ignore.xlsx"
            wb = Workbook()
            ws = wb.active
            
            # White background with black text (should be ignored by default)
            ws['A1'] = 'Normal'
            ws['A1'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            ws['A1'].font = Font(color='000000')
            
            # Colored background with colored text
            ws['B1'] = 'Colored'
            ws['B1'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            ws['B1'].font = Font(color='00FF00')
            
            wb.save(xlsx_path)
            
            converter = XlsxConverter()
            
            # Test 1: With defaults (ignore white bg and black fg)
            result = converter.convert_to_csv(
                str(xlsx_path), 'Sheet',
                include_colors=True  # Both fg and bg
            )
            lines = result.strip().split('\n')
            # Normal cell should have no colors (white bg and black fg ignored)
            assert 'Normal,' in lines[0] or ',Normal' in lines[0]
            assert '{#FFFFFF}' not in lines[0]  # White bg ignored
            assert '{fc:#000000}' not in lines[0]  # Black fg ignored
            # Colored cell should have both colors
            assert '{#FF0000}' in lines[0]  # Red bg included
            assert '{fc:#00FF00}' in lines[0]  # Green fg included
            
            # Test 2: Override with empty ignore lists
            result = converter.convert_to_csv(
                str(xlsx_path), 'Sheet',
                include_colors=True,
                ignore_bg_colors='',  # Don't ignore any bg colors
                ignore_fg_colors=''   # Don't ignore any fg colors
            )
            lines = result.strip().split('\n')
            # Now white bg and black fg should be included
            assert '{#FFFFFF}' in lines[0]  # White bg now included
            assert '{fc:#000000}' in lines[0]  # Black fg now included
            
            # Test 3: Custom ignore lists
            result = converter.convert_to_csv(
                str(xlsx_path), 'Sheet',
                include_colors=True,
                ignore_bg_colors='#FF0000',  # Ignore red bg
                ignore_fg_colors='#00FF00'   # Ignore green fg
            )
            lines = result.strip().split('\n')
            # Red bg and green fg should be ignored
            assert '{#FF0000}' not in lines[0]  # Red bg ignored
            assert '{fc:#00FF00}' not in lines[0]  # Green fg ignored
            # White bg and black fg should be included (not in ignore list)
            assert '{#FFFFFF}' in lines[0]  # White bg included
            assert '{fc:#000000}' in lines[0]  # Black fg included (with background)
    
    def test_max_rows_columns(self):
        """Test max rows and columns limits."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create test file with many rows and columns
            xlsx_path = Path(temp_dir) / "test_limits.xlsx"
            wb = Workbook()
            ws = wb.active
            
            # Create 10x10 grid
            for row in range(1, 11):
                for col in range(1, 11):
                    ws.cell(row=row, column=col, value=f"R{row}C{col}")
            
            wb.save(xlsx_path)
            
            # Convert with limits
            converter = XlsxConverter()
            result = converter.convert_to_csv(
                str(xlsx_path),
                'Sheet',
                max_rows=3,
                max_columns=4
            )
            
            # Parse result
            lines = result.strip().split('\n')
            # Should have only 3 rows
            assert len(lines) == 3
            
            # Each row should have only 4 columns
            for line in lines:
                cells = line.split(',')
                assert len(cells) == 4


if __name__ == '__main__':
    pytest.main([__file__, '-v'])