"""
Main XLSX to Flagged CSV converter implementation.
"""

from typing import Optional, Literal, Dict, Any
from pathlib import Path
import random
import warnings
import colorsys
import zipfile
import xml.etree.ElementTree as etree

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from .formatter import ExcelFormatter


class XlsxConverterConfig(BaseModel):
    """Configuration for XLSX conversion."""
    
    keep_default_na: bool = Field(
        default=False,
        description="Whether to keep default NA values in pandas"
    )
    index: bool = Field(
        default=False,
        description="Whether to include index in CSV output"
    )
    header: bool = Field(
        default=False,
        description="Whether to include header in CSV output"
    )
    keep_empty_lines: bool = Field(
        default=False,
        description="Whether to preserve empty rows to maintain original row positions"
    )
    add_location: bool = Field(
        default=False,
        description="Whether to add location coordinates {l:A5} to non-empty cells"
    )


class XlsxConverter:
    """Convert XLSX files to CSV format with optional formatting flags."""
    
    def __init__(self, config: Optional[XlsxConverterConfig] = None):
        """
        Initialize the converter.
        
        Args:
            config: Optional configuration for the converter
        """
        self.config = config or XlsxConverterConfig()
        self._patch_openpyxl_colors()
        self._cached_theme_colors = None
    
    def _patch_openpyxl_colors(self):
        """Patch openpyxl to handle color validation issues."""
        import openpyxl.styles.colors
        original_rgb_set = openpyxl.styles.colors.RGB.__set__
        
        def patched_rgb_set(self, instance, value):
            if value is not None:
                if isinstance(value, str):
                    value = ''.join(c for c in value if c in '0123456789ABCDEFabcdef')
                    if len(value) == 6:
                        value = 'FF' + value
                    elif len(value) < 6:
                        value = value.ljust(8, '0')
                    elif len(value) > 8:
                        value = value[:8]
            try:
                original_rgb_set(self, instance, value)
            except ValueError:
                instance._rgb = None
        
        openpyxl.styles.colors.RGB.__set__ = patched_rgb_set
    
    def convert_to_csv(
        self,
        input_file_path: str,
        tab_name: str,
        output_format: Literal["csv", "html", "markdown"] = "csv",
        include_colors: bool = False,
        include_bg_colors: bool = False,
        include_fg_colors: bool = False,
        signal_merge: bool = False,
        preserve_formats: bool = False,
        ignore_colors: Optional[str] = None,
        ignore_bg_colors: Optional[str] = None,
        ignore_fg_colors: Optional[str] = None,
        keep_empty_lines: Optional[bool] = None,
        add_location: Optional[bool] = None,
        max_rows: Optional[int] = None,
        max_columns: Optional[int] = None
    ) -> str:
        """
        Convert an XLSX file to CSV with optional formatting flags.
        
        Args:
            input_file_path: Path to the XLSX file
            tab_name: Name of the sheet to convert
            output_format: Format to convert to (csv, html, or markdown)
            include_colors: Whether to include both foreground and background colors
            include_bg_colors: Whether to include background colors as {bc:#RRGGBB} or {#RRGGBB} flags
            include_fg_colors: Whether to include foreground colors as {fc:#RRGGBB} flags
            signal_merge: Whether to include merge information as {MG:XXXXXX} flags
            preserve_formats: Whether to preserve cell formatting (e.g., $500 instead of 500)
            ignore_colors: Comma-separated hex colors to ignore for both fg and bg (applies defaults)
            ignore_bg_colors: Comma-separated hex colors to ignore for background (default: "#FFFFFF")
            ignore_fg_colors: Comma-separated hex colors to ignore for foreground (default: "#000000")
            keep_empty_lines: Whether to preserve empty rows (overrides config if provided)
            add_location: Whether to add location coordinates {l:A5} to cells (overrides config if provided)
            max_rows: Maximum number of rows to process (default: 300)
            max_columns: Maximum number of columns to process (default: 100)
            
        Returns:
            Formatted string in the requested format
            
        Raises:
            FileNotFoundError: If the input file doesn't exist
            ValueError: If the tab name doesn't exist in the XLSX file
        """
        # Verify file exists
        path = Path(input_file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {input_file_path}")
        if path.suffix.lower() not in ['.xlsx', '.xls']:
            raise ValueError(f"File must be an Excel file (xlsx/xls), got: {path.suffix}")
        
        try:
            # Determine feature flags
            should_keep_empty_lines = keep_empty_lines if keep_empty_lines is not None else self.config.keep_empty_lines
            should_add_location = add_location if add_location is not None else self.config.add_location
            
            # Set default max rows and columns
            max_rows = max_rows if max_rows is not None else 300
            max_columns = max_columns if max_columns is not None else 100
            
            # Handle include_colors as master flag for both fg and bg
            if include_colors:
                include_bg_colors = True
                include_fg_colors = True
            
            # Process ignore color lists with defaults
            bg_ignore_list = set()
            fg_ignore_list = set()
            
            # If ignore_colors is specified, it overrides individual lists
            if ignore_colors is not None:
                # Parse the combined ignore list
                colors = [c.strip().upper().replace('#', '') for c in ignore_colors.split(',') if c.strip()]
                bg_ignore_list.update(colors)
                fg_ignore_list.update(colors)
                # Also apply defaults when using ignore_colors
                bg_ignore_list.add('FFFFFF')  # Default: ignore white backgrounds
                fg_ignore_list.add('000000')  # Default: ignore black foreground
            else:
                # Use individual ignore lists or defaults
                if ignore_bg_colors is not None:
                    bg_ignore_list.update(c.strip().upper().replace('#', '') for c in ignore_bg_colors.split(',') if c.strip())
                else:
                    bg_ignore_list.add('FFFFFF')  # Default: ignore white backgrounds
                
                if ignore_fg_colors is not None:
                    fg_ignore_list.update(c.strip().upper().replace('#', '') for c in ignore_fg_colors.split(',') if c.strip())
                else:
                    fg_ignore_list.add('000000')  # Default: ignore black foreground
            
            # Determine whether to use special formatting
            if include_bg_colors or include_fg_colors or signal_merge or preserve_formats or should_keep_empty_lines or should_add_location:
                df = self._read_excel_with_formatting(
                    input_file_path, tab_name, 
                    include_bg_colors, include_fg_colors,
                    signal_merge, preserve_formats, 
                    bg_ignore_list, fg_ignore_list,
                    should_keep_empty_lines, should_add_location,
                    max_rows, max_columns
                )
            else:
                df = self._read_excel_with_fallback(
                    input_file_path, tab_name, self.config.keep_default_na,
                    max_rows, max_columns, should_keep_empty_lines
                )
            
            # Convert to requested format
            if output_format == "csv":
                return df.to_csv(index=self.config.index, header=self.config.header)
            elif output_format == "html":
                return df.to_html(index=self.config.index, header=self.config.header)
            elif output_format == "markdown":
                return df.to_markdown(index=self.config.index)
            
        except ValueError as e:
            if "No sheet named" in str(e) or "not found" in str(e):
                raise ValueError(f"Sheet '{tab_name}' not found in {path.name}")
            raise
    
    def _read_excel_with_fallback(self, file_path: str, sheet_name: str, keep_default_na: bool, max_rows: int = 300, max_columns: int = 100, keep_empty_lines: bool = False) -> pd.DataFrame:
        """Read Excel file with multiple engine fallbacks.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to read
            keep_default_na: Whether to keep default NA values
            max_rows: Maximum number of rows to read
            max_columns: Maximum number of columns to read
            keep_empty_lines: Whether to keep empty rows
            
        Returns:
            pd.DataFrame: The loaded dataframe
        """
        exceptions = []
        
        # Try calamine engine first
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, 
                               keep_default_na=keep_default_na, engine='calamine', nrows=max_rows, header=None)
            # Limit columns
            if len(df.columns) > max_columns:
                df = df.iloc[:, :max_columns]
            # Rename columns to Excel-style letters
            df.columns = [get_column_letter(i + 1) for i in range(len(df.columns))]
            # Remove empty rows if needed
            if not keep_empty_lines:
                df = self._remove_empty_rows(df)
            # Always trim trailing empty rows
            df = self._trim_trailing_empty_rows(df)
            return df
        except Exception as e:
            exceptions.append(f"Calamine: {e}")
        
        # Try openpyxl
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, 
                               keep_default_na=keep_default_na, engine='openpyxl', nrows=max_rows, header=None)
            # Limit columns
            if len(df.columns) > max_columns:
                df = df.iloc[:, :max_columns]
            # Rename columns to Excel-style letters
            df.columns = [get_column_letter(i + 1) for i in range(len(df.columns))]
            # Remove empty rows if needed
            if not keep_empty_lines:
                df = self._remove_empty_rows(df)
            # Always trim trailing empty rows
            df = self._trim_trailing_empty_rows(df)
            return df
        except Exception as e:
            exceptions.append(f"Openpyxl: {e}")
        
        # Try xlrd
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, 
                               keep_default_na=keep_default_na, engine='xlrd', nrows=max_rows, header=None)
            # Limit columns
            if len(df.columns) > max_columns:
                df = df.iloc[:, :max_columns]
            # Rename columns to Excel-style letters
            df.columns = [get_column_letter(i + 1) for i in range(len(df.columns))]
            # Remove empty rows if needed
            if not keep_empty_lines:
                df = self._remove_empty_rows(df)
            # Always trim trailing empty rows
            df = self._trim_trailing_empty_rows(df)
            return df
        except Exception as e:
            exceptions.append(f"Xlrd: {e}")
        
        # Direct openpyxl reading
        try:
            warnings.filterwarnings('ignore')
            wb = load_workbook(file_path, data_only=True, keep_links=False)
            ws = wb[sheet_name]
            
            data = []
            for row in ws.iter_rows(values_only=True, max_row=max_rows, max_col=max_columns):
                data.append(row)
            
            if data:
                df = pd.DataFrame(data)
                # Rename columns to Excel-style letters
                df.columns = [get_column_letter(i + 1) for i in range(len(df.columns))]
                if not keep_default_na:
                    df = df.fillna('')
                # Remove empty rows if needed
                if not keep_empty_lines:
                    df = self._remove_empty_rows(df)
                # Always trim trailing empty rows
                df = self._trim_trailing_empty_rows(df)
                return df
            return pd.DataFrame()
            
        except Exception as e:
            exceptions.append(f"Direct openpyxl: {e}")
        
        # Check for sheet not found errors
        for exc in exceptions:
            if "not found" in exc.lower() or "does not exist" in exc.lower():
                raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file")
        
        raise Exception(f"Unable to read Excel file: {'; '.join(exceptions)}")
    
    def _extract_theme_colors(self, file_path: str) -> Dict[int, str]:
        """Extract theme colors from XLSX file."""
        theme_colors = {}
        
        try:
            with zipfile.ZipFile(file_path, 'r') as zip_file:
                if 'xl/theme/theme1.xml' in zip_file.namelist():
                    theme_xml = zip_file.read('xl/theme/theme1.xml')
                    root = etree.fromstring(theme_xml)
                    
                    ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
                    color_scheme = root.find('.//a:clrScheme', ns)
                    
                    if color_scheme is not None:
                        color_map = [
                            ('lt1', 0), ('dk1', 1), ('lt2', 2), ('dk2', 3),
                            ('accent1', 4), ('accent2', 5), ('accent3', 6),
                            ('accent4', 7), ('accent5', 8), ('accent6', 9),
                            ('hlink', 10), ('folHlink', 11)
                        ]
                        
                        for color_name, idx in color_map:
                            elem = color_scheme.find(f'.//a:{color_name}', ns)
                            if elem is not None:
                                srgb = elem.find('.//a:srgbClr', ns)
                                sys_color = elem.find('.//a:sysClr', ns)
                                
                                if srgb is not None:
                                    theme_colors[idx] = srgb.get('val')
                                elif sys_color is not None:
                                    theme_colors[idx] = sys_color.get('lastClr', '000000')
        except Exception:
            pass
        
        return theme_colors
    
    def _apply_tint(self, rgb_hex: str, tint: float) -> str:
        """Apply tint to a color."""
        if not tint or tint == 0:
            return rgb_hex
        
        # Convert hex to RGB
        r = int(rgb_hex[0:2], 16) / 255.0
        g = int(rgb_hex[2:4], 16) / 255.0
        b = int(rgb_hex[4:6], 16) / 255.0
        
        # Convert to HSL
        h, l, s = colorsys.rgb_to_hls(r, g, b)
        
        # Apply tint
        if tint < 0:
            l = l * (1 + tint)  # Make darker
        else:
            l = l + (1 - l) * tint  # Make lighter
        
        # Convert back to RGB
        r, g, b = colorsys.hls_to_rgb(h, l, s)
        
        return f"{int(r * 255):02X}{int(g * 255):02X}{int(b * 255):02X}"
    
    def _read_excel_with_formatting(
        self,
        file_path: str,
        sheet_name: str,
        include_bg_colors: bool = False,
        include_fg_colors: bool = False,
        signal_merge: bool = False,
        preserve_formats: bool = False,
        bg_ignore_list: set = None,
        fg_ignore_list: set = None,
        keep_empty_lines: bool = False,
        add_location: bool = False,
        max_rows: int = 300,
        max_columns: int = 100
    ) -> pd.DataFrame:
        """Read Excel file and extract formatting information.
        
        This implementation preserves the exact row/column structure and appends
        formatting information directly to cell values using:
        - Colors: value{#color}
        - Merged cells: value{MG:xxxxx} where xxxxx is a unique identifier
        - Location: value{l:A5} where A5 is the Excel coordinate
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to read
            include_bg_colors: Whether to include background color information
            include_fg_colors: Whether to include foreground color information
            signal_merge: Whether to include merge information
            preserve_formats: Whether to preserve cell formatting
            bg_ignore_list: Set of background colors to ignore (hex without #)
            fg_ignore_list: Set of foreground colors to ignore (hex without #)
            keep_empty_lines: Whether to preserve empty rows to maintain original row positions
            add_location: Whether to add location coordinates to non-empty cells
            max_rows: Maximum number of rows to process
            max_columns: Maximum number of columns to process
            
        Returns:
            pd.DataFrame: DataFrame with original structure and formatting embedded in values
        """
        # Use provided ignore lists or empty sets
        bg_ignore_list = bg_ignore_list or set()
        fg_ignore_list = fg_ignore_list or set()
        
        wb = load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file")
        
        ws = wb[sheet_name]
        
        # Clear cached theme colors
        self._cached_theme_colors = None
        
        # Build merge map
        merge_map = {}
        if signal_merge and ws.merged_cells:
            for merged_range in ws.merged_cells.ranges:
                # Only process merged cells within our limits
                if merged_range.min_row <= max_rows and merged_range.min_col <= max_columns:
                    merge_id = str(random.randint(100000, 999999))
                    # Mark all cells in the merged range with the same ID (within limits)
                    for row in range(merged_range.min_row, min(merged_range.max_row + 1, max_rows + 1)):
                        for col in range(merged_range.min_col, min(merged_range.max_col + 1, max_columns + 1)):
                            merge_map[(row, col)] = merge_id
        
        # Process all cells
        processed_data = []
        
        for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, max_col=max_columns), 1):
            row_data = []
            has_content = False
            
            for col_idx, cell in enumerate(row, 1):
                value = cell.value
                if value is not None:
                    has_content = True
                
                # Apply formatting if requested
                if preserve_formats and value is not None:
                    try:
                        if cell.number_format and cell.number_format != 'General':
                            value = ExcelFormatter.format_value(value, cell.number_format)
                    except:
                        pass
                
                formatting_parts = []
                
                # Extract background color if requested
                has_background = False
                if include_bg_colors:
                    bg_color_hex = self._extract_cell_bg_color(cell, file_path)
                    if bg_color_hex:
                        color_to_check = bg_color_hex[1:] if bg_color_hex.startswith('#') else bg_color_hex
                        if color_to_check.upper() not in bg_ignore_list:
                            # Use backward-compatible {#RRGGBB} format by default
                            formatting_parts.append(f"{{{bg_color_hex}}}")
                            has_background = True
                
                # Extract foreground color if requested and cell has content
                if include_fg_colors and value is not None:
                    fg_color_hex = self._extract_cell_fg_color(cell, file_path)
                    if fg_color_hex:
                        color_to_check = fg_color_hex[1:] if fg_color_hex.startswith('#') else fg_color_hex
                        if color_to_check.upper() not in fg_ignore_list:
                            # Only include black foreground color if there's also a background color (for contrast)
                            if color_to_check.upper() == '000000':
                                # Check if this cell has a background color
                                if not has_background and not include_bg_colors:
                                    # Need to check for background if we haven't already
                                    bg_color = self._extract_cell_bg_color(cell, file_path)
                                    has_background = bg_color is not None
                                # Only include black text if there's a background
                                if has_background:
                                    formatting_parts.append(f"{{fc:{fg_color_hex}}}")
                            else:
                                # Non-black colors are always included
                                formatting_parts.append(f"{{fc:{fg_color_hex}}}")
                
                # Add merge info if requested
                if signal_merge and (cell.row, cell.column) in merge_map:
                    merge_id = merge_map[(cell.row, cell.column)]
                    formatting_parts.append(f"{{MG:{merge_id}}}")
                
                # Add location if requested and cell is non-empty (has value, color, or merge flag)
                if add_location and (value is not None or formatting_parts):
                    column_letter = get_column_letter(cell.column)
                    location_tag = f"{{l:{column_letter}{cell.row}}}"
                    formatting_parts.append(location_tag)
                
                # Combine value and formatting
                if value is not None:
                    cell_value = str(value) + ''.join(formatting_parts)
                elif formatting_parts:
                    cell_value = ''.join(formatting_parts)
                else:
                    cell_value = None
                
                row_data.append(cell_value)
            
            # keep_empty_lines is useful if we want to preserve the original row positions
            if has_content or keep_empty_lines:
                processed_data.append(row_data)
        
        if not processed_data:
            return pd.DataFrame()
        
        # Create DataFrame
        # Ensure all rows have same number of columns
        max_cols = max(len(row) for row in processed_data) if processed_data else 0
        for row in processed_data:
            while len(row) < max_cols:
                row.append(None)
        
        # Create DataFrame with column letters as headers (A, B, C, ...)
        # This provides meaningful column names for CSV output
        column_names = [get_column_letter(i) for i in range(1, max_cols + 1)]
        df = pd.DataFrame(processed_data, columns=column_names)
        
        # Fill NaN values if needed
        if not self.config.keep_default_na:
            df = df.fillna('')
        
        # Always trim trailing empty rows (even when keep_empty_lines=True)
        # keep_empty_lines only preserves empty rows within the content area
        df = self._trim_trailing_empty_rows(df)
        
        return df
    
    def _extract_cell_fg_color(self, cell, file_path: str) -> Optional[str]:
        """Extract foreground (font) color from a cell."""
        if not cell.font or not cell.font.color:
            return None
        
        color = cell.font.color
        
        try:
            if hasattr(color, 'type'):
                if color.type == 'theme' and hasattr(color, 'theme') and color.theme is not None:
                    # Load theme colors
                    if self._cached_theme_colors is None:
                        self._cached_theme_colors = self._extract_theme_colors(file_path)
                    
                    # Default theme colors if extraction fails
                    if not self._cached_theme_colors:
                        self._cached_theme_colors = {
                            0: "FFFFFF", 1: "000000", 2: "E7E6E6", 3: "44546A",
                            4: "5B9BD5", 5: "ED7D31", 6: "A5A5A5", 7: "FFC000",
                            8: "4472C4", 9: "70AD47"
                        }
                    
                    base_color = self._cached_theme_colors.get(color.theme, "000000")
                    
                    # Apply tint if present
                    tint = getattr(color, 'tint', 0)
                    if tint and tint != 0:
                        final_color = self._apply_tint(base_color, tint)
                        return f"#{final_color}"
                    return f"#{base_color}"
                
                elif color.type == 'rgb' and hasattr(color, 'rgb'):
                    rgb = color.rgb
                    if rgb and isinstance(rgb, str) and len(rgb) >= 6:
                        # Extract last 6 characters (ignore alpha channel if present)
                        color_hex = rgb[-6:]
                        return f"#{color_hex}"
            
            # Try direct RGB attribute
            if hasattr(color, 'rgb') and color.rgb:
                rgb = color.rgb
                if isinstance(rgb, str) and len(rgb) >= 6:
                    color_hex = rgb[-6:]
                    return f"#{color_hex}"
        except:
            pass
        
        return None
    
    def _extract_cell_bg_color(self, cell, file_path: str) -> Optional[str]:
        """Extract background color from a cell."""
        if not cell.fill or cell.fill.patternType != 'solid':
            return None
        
        color = cell.fill.fgColor or cell.fill.start_color
        if not color:
            return None
        
        try:
            if hasattr(color, 'type'):
                if color.type == 'theme' and hasattr(color, 'theme') and color.theme is not None:
                    # Load theme colors
                    if self._cached_theme_colors is None:
                        self._cached_theme_colors = self._extract_theme_colors(file_path)
                    
                    # Default theme colors if extraction fails
                    if not self._cached_theme_colors:
                        self._cached_theme_colors = {
                            0: "FFFFFF", 1: "000000", 2: "E7E6E6", 3: "44546A",
                            4: "5B9BD5", 5: "ED7D31", 6: "A5A5A5", 7: "FFC000",
                            8: "4472C4", 9: "70AD47"
                        }
                    
                    base_color = self._cached_theme_colors.get(color.theme, "000000")
                    
                    # Apply tint if present
                    tint = getattr(color, 'tint', 0)
                    if tint and tint != 0:
                        final_color = self._apply_tint(base_color, tint)
                        return f"#{final_color}"
                    return f"#{base_color}"
                
                elif color.type == 'indexed' and hasattr(color, 'indexed'):
                    # Try to get indexed color from workbook or use openpyxl's COLOR_INDEX
                    idx = color.indexed
                    if idx is not None:
                        # First try to get from workbook's custom colors if available
                        wb = cell.parent.parent  # cell -> worksheet -> workbook
                        if hasattr(wb, '_colors') and wb._colors and idx < len(wb._colors):
                            rgb = wb._colors[idx]
                            # Remove alpha channel if present (first 2 chars)
                            if len(rgb) == 8 and rgb.startswith('00'):
                                return f"#{rgb[2:]}"
                            elif len(rgb) == 6:
                                return f"#{rgb}"
                        
                        # Fall back to openpyxl's default COLOR_INDEX
                        try:
                            from openpyxl.styles.colors import COLOR_INDEX
                            if idx < len(COLOR_INDEX):
                                rgb = COLOR_INDEX[idx]
                                # Remove alpha channel if present
                                if len(rgb) == 8:
                                    return f"#{rgb[2:]}"
                                elif len(rgb) == 6:
                                    return f"#{rgb}"
                        except ImportError:
                            pass
                    
                    # If all else fails, return white (most common for undefined indexed colors)
                    return "#FFFFFF"
                
                elif color.type == 'rgb' or not hasattr(color, 'type'):
                    if hasattr(color, 'rgb'):
                        rgb = color.rgb
                        if isinstance(rgb, str) and 'Values must be' not in rgb:
                            if len(rgb) == 8:
                                return f"#{rgb[2:]}"
                            elif len(rgb) == 6:
                                return f"#{rgb}"
        except:
            pass
        
        return None
    
    def _remove_empty_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        """Remove rows that are entirely empty or contain only empty strings.
        
        Args:
            df: DataFrame to filter
            
        Returns:
            pd.DataFrame: DataFrame with empty rows removed
        """
        if df.empty:
            return df
        
        # Filter out rows where all values are either NaN or empty strings
        mask = df.apply(lambda row: any(pd.notna(val) and str(val).strip() != '' for val in row), axis=1)
        return df[mask]
    
    def _trim_trailing_empty_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        """Trim trailing empty rows and columns from DataFrame.
        
        This is always applied, even when keep_empty_lines=True, because
        keep_empty_lines is meant to preserve empty rows within the content area,
        not trailing empty rows.
        
        Args:
            df: DataFrame to trim
            
        Returns:
            pd.DataFrame: DataFrame with trailing empty rows and columns removed
        """
        if df.empty:
            return df
        
        # Find the last non-empty row by checking from the end
        last_non_empty_row = -1
        for idx in range(len(df) - 1, -1, -1):
            row = df.iloc[idx]
            # Check if row has any non-empty values
            has_content = False
            for val in row:
                if pd.notna(val) and str(val).strip() != '':
                    has_content = True
                    break
            
            if has_content:
                last_non_empty_row = idx
                break
        
        # If all rows are empty, return empty DataFrame
        if last_non_empty_row == -1:
            return pd.DataFrame(columns=df.columns)
        
        # Trim to last non-empty row
        df = df.iloc[:last_non_empty_row + 1]
        
        # Also find the last non-empty column
        last_non_empty_col = -1
        for col_idx in range(len(df.columns) - 1, -1, -1):
            col = df.iloc[:, col_idx]
            # Check if column has any non-empty values
            has_content = False
            for val in col:
                if pd.notna(val) and str(val).strip() != '':
                    has_content = True
                    break
            
            if has_content:
                last_non_empty_col = col_idx
                break
        
        # If all columns are empty, return DataFrame with no columns
        if last_non_empty_col == -1:
            return pd.DataFrame()
        
        # Return DataFrame trimmed to last non-empty column
        return df.iloc[:, :last_non_empty_col + 1]